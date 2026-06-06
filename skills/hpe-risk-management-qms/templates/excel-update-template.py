#!/usr/bin/env python3
"""
Starter template for future QMS Risk Register updates.
Copy this file and adapt the UPDATES dict for the next risk item.
Includes Impact (H), Likelihood (I), Level (J) — always compute Level = Impact * Likelihood.
"""
import sys
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
GREEN_FONT = Font(color="006100")
WRAP = Alignment(wrapText=True, vertical="top")

def update_excel_cells(wb_path: str, sheet_name: str, updates: dict) -> None:
    wb = load_workbook(wb_path, read_only=False, data_only=False)
    ws = wb[sheet_name]
    # Auto-compute Level (J) if Impact (H) and Likelihood (I) are being set
    if f"H{ROW}" in updates and f"I{ROW}" in updates:
        imp = updates[f"H{ROW}"]
        like = updates[f"I{ROW}"]
        if isinstance(imp, int) and isinstance(like, int):
            updates[f"J{ROW}"] = imp * like
    for cell_ref, new_value in updates.items():
        cell = ws[cell_ref]
        cell.value = new_value
        cell.fill = GREEN_FILL
        cell.font = GREEN_FONT
        cell.alignment = WRAP
    wb.save(wb_path)
    print(f"Updated {len(updates)} cells in sheet '{sheet_name}'.")
    for ref, val in updates.items():
        txt = str(val)[:80]
        print(f"  {ref}: {txt}{'...' if len(txt)>80 else ''}")

if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Usage: python excel-update-template.py <path-to-risk-register-xlsx>")
        sys.exit(1)
    SHEET = "Risk Assessment"
    ROW = 3  # Adjust per risk row
    UPDATES = {
        f"H{ROW}": 3,         # Impact (1-4)
        f"I{ROW}": 2,         # Likelihood (1-4)
        # J{ROW} auto-computed
        f"K{ROW}": "Your mitigation plan here …",
        f"L{ROW}": "Your contingency plan here …",
        f"P{ROW}": "Your risk response decision here …",
        f"Q{ROW}": "To be determined",
        f"R{ROW}": "In Treatment",
    }
    update_excel_cells(sys.argv[1], SHEET, UPDATES)